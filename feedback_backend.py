from __future__ import annotations

import os
import sqlite3
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path
from statistics import mean
import calendar
from typing import List

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "feedback.db"
EXCEL_PATH = BASE_DIR / "feedback_data.xlsx"

app = FastAPI(title="Feedback Backend", version="2.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

POSITIVE_WORDS = {
    "good", "great", "excellent", "amazing", "fast", "friendly", "clean", "helpful",
    "respectful", "professional", "organized", "smooth", "quick", "nice", "love",
    "clear", "comfortable", "supportive", "efficient",
}
NEGATIVE_WORDS = {
    "bad", "slow", "late", "waiting", "delay", "crowded", "dirty", "rude", "problem",
    "issue", "complaint", "poor", "worse", "worst", "confusing", "long", "tired",
    "noise", "unclear", "ignored", "frustrating",
}
KEYWORDS = {
    "waiting time": ["wait", "waiting", "delay", "queue", "line", "time"],
    "staff professionalism": ["staff", "professional", "respectful", "friendly", "nurse", "doctor"],
    "cleanliness": ["clean", "dirty", "hygiene"],
    "communication": ["communication", "clear", "explained", "update", "told"],
    "speed": ["fast", "slow", "quick", "response", "service"],
}
SEED_ROWS = [
    ("Emergency Department", "Triage Team", 5, 4, 4, "The nurse was respectful and quick, but the waiting time was still long."),
    ("Emergency Department", "Trauma Team", 4, 4, 5, "Professional team and clear communication during treatment."),
    ("Outpatient Clinics", "General Clinics Team", 5, 5, 5, "Excellent organization and very friendly staff."),
    ("Outpatient Clinics", "Specialist Clinics Team", 3, 4, 3, "Doctor was good but the queue updates were unclear."),
    ("Inpatient Wards", "Patient Care Team", 4, 5, 4, "Clean ward and supportive staff."),
    ("Intensive Care Unit (ICU)", "Monitoring Team", 5, 5, 5, "Amazing care and professional communication."),
    ("Operating Rooms", "Recovery Room Team", 4, 4, 4, "Good service overall and comfortable process."),
    ("Laboratory", "Laboratory Reception", 2, 3, 2, "The line was long and response was slow."),
    ("Radiology", "X-Ray Unit", 4, 3, 4, "Helpful team but there was a delay before imaging."),
    ("Pharmacy", "Dispensing Counter", 5, 4, 4, "Fast pickup and clear instructions from the pharmacist."),
    ("Physiotherapy", "Therapy Sessions", 5, 5, 4, "Very professional therapist and smooth session."),
    ("Nutrition Services", "Dietitian Unit", 4, 5, 4, "Clear explanation and respectful communication."),
    ("Administration", "Front Desk", 3, 3, 2, "The process was confusing and took too long."),
    ("Human Resources (HR)", "Employee Support", 4, 4, 5, "Helpful team and quick response."),
    ("Finance Department", "Billing", 2, 2, 3, "Billing queue was crowded and the explanation was not clear enough."),
    ("Information Technology (IT)", "Support Desk", 5, 5, 5, "Issue solved quickly and staff were professional."),
    ("Patient Relations", "Complaints Unit", 4, 4, 4, "They listened carefully and handled my concern respectfully."),
    ("Pediatrics", "Nursing Support Team", 5, 4, 5, "Very friendly and caring staff with clean rooms."),
    ("Cardiology", "Monitoring Unit", 4, 5, 4, "Good attention and clear updates from the team."),
    ("General Surgery", "Post-Operative Care Team", 3, 4, 3, "Service quality was good but waiting for discharge took a long time."),
    ("Emergency Department", "Ambulance & Transfer Team", 4, 4, 4, "Transfer was organized and staff were calm."),
    ("Outpatient Clinics", "Appointment Scheduling Team", 2, 3, 2, "Booking process was slow and confusing."),
    ("Inpatient Wards", "Ward Support Team", 5, 4, 5, "Clean environment and very kind staff."),
    ("Operating Rooms", "Sterilization Team", 5, 5, 4, "Professional process and confidence in cleanliness."),
    ("Laboratory", "Sample Collection", 3, 4, 3, "Staff were nice but there was still a long wait."),
    ("Pharmacy", "Counseling Desk", 5, 5, 5, "Excellent explanation of medications and fast service."),
    ("Administration", "Records Office", 2, 2, 2, "Poor communication and too much delay in paperwork."),
    ("Patient Relations", "Survey Support", 4, 5, 4, "Friendly and helpful support when I had a complaint."),
]

class FeedbackIn(BaseModel):
    department: str = Field(default="")
    section: str = Field(default="")
    rating1: int = Field(ge=0, le=5, default=0)
    rating2: int = Field(ge=0, le=5, default=0)
    rating3: int = Field(ge=0, le=5, default=0)
    comment: str = Field(default="")


class ChatIn(BaseModel):
    message: str = Field(min_length=1)


def contains_arabic(text: str) -> bool:
    return any("؀" <= ch <= "ۿ" for ch in (text or ""))


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def detect_sentiment(comment: str, avg_rating: float) -> str:
    text = (comment or "").lower()
    pos = sum(word in text for word in POSITIVE_WORDS)
    neg = sum(word in text for word in NEGATIVE_WORDS)
    if avg_rating >= 4 or pos > neg:
        return "Positive"
    if avg_rating <= 2 or neg > pos:
        return "Negative"
    return "Neutral"


def detect_main_topic(comment: str) -> str:
    text = (comment or "").lower()
    scores = {topic: sum(word in text for word in words) for topic, words in KEYWORDS.items()}
    best_topic, best_score = max(scores.items(), key=lambda item: item[1], default=("general", 0))
    return best_topic if best_score > 0 else "general"


def init_db() -> None:
    with get_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS feedback (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                department TEXT NOT NULL,
                section TEXT NOT NULL,
                rating1 INTEGER NOT NULL,
                rating2 INTEGER NOT NULL,
                rating3 INTEGER NOT NULL,
                average_rating REAL NOT NULL,
                comment TEXT NOT NULL
            )
            """
        )
        count = conn.execute("SELECT COUNT(*) FROM feedback").fetchone()[0]
        if count == 0:
            now = datetime.now()
            for idx, row in enumerate(SEED_ROWS):
                department, section, rating1, rating2, rating3, comment = row
                created_at = (now - timedelta(days=(len(SEED_ROWS) - idx), hours=idx % 8)).strftime("%Y-%m-%d %H:%M:%S")
                avg = round((rating1 + rating2 + rating3) / 3, 2)
                conn.execute(
                    "INSERT INTO feedback (created_at, department, section, rating1, rating2, rating3, average_rating, comment) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (created_at, department, section, rating1, rating2, rating3, avg, comment),
                )
        conn.commit()


def export_excel() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback"
    headers = ["ID", "Created At", "Department", "Section", "Rating 1", "Rating 2", "Rating 3", "Average Rating", "Comment", "Sentiment", "Main Topic"]
    ws.append(headers)
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM feedback ORDER BY id DESC").fetchall()
    for row in rows:
        comment = row["comment"] or ""
        ws.append([
            row["id"], row["created_at"], row["department"], row["section"], row["rating1"], row["rating2"], row["rating3"], round(row["average_rating"], 2),
            comment, detect_sentiment(comment, row["average_rating"]), detect_main_topic(comment)
        ])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    widths = {1: 8, 2: 22, 3: 24, 4: 24, 5: 10, 6: 10, 7: 10, 8: 14, 9: 58, 10: 14, 11: 24}
    for col_idx, width in widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)


def fetch_rows(limit: int | None = None) -> List[sqlite3.Row]:
    query = "SELECT * FROM feedback ORDER BY id DESC"
    if limit:
        query += f" LIMIT {int(limit)}"
    with get_conn() as conn:
        return conn.execute(query).fetchall()


def compute_star_counts(rows: List[sqlite3.Row]) -> List[int]:
    stars = [0, 0, 0, 0, 0]
    for row in rows:
        rounded = max(1, min(5, int(round(row["average_rating"]))))
        stars[5 - rounded] += 1
    return stars




def sentiment_class(sentiment: str) -> str:
    return sentiment.lower()


def relative_time_label(created_at_str: str) -> str:
    try:
        created_at = datetime.strptime(created_at_str, "%Y-%m-%d %H:%M:%S")
    except Exception:
        return created_at_str
    diff = datetime.now() - created_at
    if diff.days <= 0:
        hours = max(1, diff.seconds // 3600)
        if hours < 24:
            return f"{hours} hour{'s' if hours != 1 else ''} ago"
        return "Today"
    if diff.days == 1:
        return "Yesterday"
    if diff.days < 7:
        return f"{diff.days} days ago"
    return created_at.strftime("%d %b %Y")


def build_summary_payload(rows: List[sqlite3.Row]) -> dict:
    if not rows:
        return {
            "total": 0,
            "average_rating": 0,
            "positive": 0,
            "neutral": 0,
            "negative": 0,
            "satisfaction_rate": 0,
            "main_concern": "No issue yet",
            "top_positive": "No data yet",
            "best_next_step": "Collect more feedback first.",
            "latest_comments": [],
            "trend": {"labels": [], "responses": [], "satisfaction": []},
        }

    sentiments = [detect_sentiment(row["comment"], row["average_rating"]) for row in rows]
    avg_rating = round(mean([row["average_rating"] for row in rows]), 2)
    satisfaction_rate = round(((sentiments.count("Positive") + sentiments.count("Neutral")) / len(rows)) * 100)

    topic_counter = Counter(detect_main_topic(row["comment"]) for row in rows if row["comment"])
    positive_counter = Counter(detect_main_topic(row["comment"]) for row in rows if detect_sentiment(row["comment"], row["average_rating"]) == "Positive" and row["comment"])
    main_concern = topic_counter.most_common(1)[0][0].title() if topic_counter else "General"
    top_positive = positive_counter.most_common(1)[0][0].title() if positive_counter else "Professional Service"

    next_step_map = {
        "Waiting Time": "Reduce peak-hour waiting time, then monitor low ratings weekly.",
        "Communication": "Improve queue updates and clearer explanations for visitors.",
        "Cleanliness": "Review cleanliness checks during busy shifts.",
        "Speed": "Track service speed by department and fix the slowest stage first.",
        "Staff Professionalism": "Maintain current staff behavior and turn it into a repeatable standard.",
        "General": "Review repeated comments and focus on the lowest-rated department first.",
    }
    best_next_step = next_step_map.get(main_concern, next_step_map["General"])

    monthly = {}
    for row in rows:
        try:
            dt = datetime.strptime(row["created_at"], "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
        key = (dt.year, dt.month)
        monthly.setdefault(key, []).append(row)
    last_keys = sorted(monthly.keys())[-6:]
    trend = {
        "labels": [calendar.month_abbr[m] for _, m in last_keys],
        "responses": [len(monthly[k]) for k in last_keys],
        "satisfaction": [round(mean(r["average_rating"] for r in monthly[k]) * 20) for k in last_keys],
    }

    latest_comments = []
    for row in rows[:20]:
        sentiment = detect_sentiment(row["comment"], row["average_rating"])
        latest_comments.append({
            "department": row["department"] or "General",
            "section": row["section"] or "General",
            "comment": row["comment"] or "No comment provided.",
            "score": f"{int(round(row['average_rating']))}/5",
            "sentiment": sentiment,
            "sentiment_class": sentiment_class(sentiment),
            "time_label": relative_time_label(row["created_at"]),
        })

    return {
        "total": len(rows),
        "average_rating": avg_rating,
        "positive": sentiments.count("Positive"),
        "neutral": sentiments.count("Neutral"),
        "negative": sentiments.count("Negative"),
        "satisfaction_rate": satisfaction_rate,
        "main_concern": main_concern,
        "top_positive": top_positive,
        "best_next_step": best_next_step,
        "latest_comments": latest_comments,
        "trend": trend,
    }
def build_feedback_context(rows: List[sqlite3.Row]) -> str:
    if not rows:
        return "No feedback available yet."
    sentiments = Counter(detect_sentiment(row["comment"], row["average_rating"]) for row in rows)
    topics = Counter(detect_main_topic(row["comment"]) for row in rows if row["comment"])
    departments = Counter(row["department"] for row in rows)
    avg = mean(row["average_rating"] for row in rows)
    latest_comments = [f"- {row['department']} / {row['section']}: {row['comment']}" for row in rows[:12] if (row['comment'] or '').strip()]
    return (
        f"Total responses: {len(rows)}\n"
        f"Average rating: {avg:.2f}/5\n"
        f"Sentiments: {dict(sentiments)}\n"
        f"Top topics: {topics.most_common(5)}\n"
        f"Top departments: {departments.most_common(5)}\n"
        "Recent comments:\n" + "\n".join(latest_comments)
    )


def is_greeting_message(user_message: str) -> bool:
    text = (user_message or "").strip().lower()
    greeting_set = {
        "hi", "hello", "hey", "hey there", "good morning", "good afternoon", "good evening",
        "مرحبا", "اهلا", "أهلا", "هلا", "السلام عليكم", "سلام", "هاي"
    }
    return text in greeting_set


def is_feedback_related(user_message: str) -> bool:
    text = (user_message or "").lower()
    keywords = [
        "feedback", "dashboard", "comment", "comments", "rating", "ratings", "survey", "surveys",
        "department", "departments", "sentiment", "satisfaction", "trend", "trends", "chart", "charts",
        "manager", "response", "responses", "issue", "issues", "complaint", "complaints",
        "hospital", "section", "summary", "anonymous comments", "kpi", "score", "scores",
        "ملاحظات", "استبيان", "الداشبورد", "تعليق", "تعليقات", "تقييم", "التقييم", "قسم", "الأقسام",
        "شارت", "رسم", "رضا", "المشروع", "نتائج", "النتايج", "تحليل"
    ]
    return any(k in text for k in keywords)


def fallback_ai_reply(user_message: str, rows: List[sqlite3.Row]) -> str:
    is_ar = contains_arabic(user_message)

    if is_greeting_message(user_message):
        if is_ar:
            return "أهلًا! أنا جاهز أساعدك في تحليل نتائج الملاحظات، التعليقات، التقييمات، الأقسام، والاتجاهات العامة بشكل طبيعي وواضح."
        return "Hi! I'm ready to help you analyze feedback results, comments, ratings, departments, and trends in a natural way."

    if not is_feedback_related(user_message):
        if is_ar:
            return "أقدر أساعدك هنا بشكل طبيعي، لكن دوري في هذه الصفحة يركز على مشروع الملاحظات فقط. اسألني عن التعليقات، التقييمات، الأقسام، اتجاهات الرضا، أو أفضل خطوة تالية للإدارة."
        return "I can chat naturally here, but on this page I'm focused on the feedback project only. Ask me about comments, ratings, departments, satisfaction trends, or the best next step for management."

    if not rows:
        if is_ar:
            return "لا توجد ملاحظات محفوظة حتى الآن. بمجرد وصول ردود جديدة أقدر ألخص لك الاتجاهات وأبرز الملاحظات والخطوات المقترحة."
        return "No feedback has been submitted yet. Once new responses arrive, I can summarize the trends, key comments, and recommended next steps."

    avg = mean([row["average_rating"] for row in rows])
    sentiments = Counter(detect_sentiment(row["comment"], row["average_rating"]) for row in rows)
    topic = Counter(detect_main_topic(row["comment"]) for row in rows if row["comment"]).most_common(1)
    top_topic = topic[0][0] if topic else "general"
    recent = [row["comment"] for row in rows if row["comment"]][:2]

    if is_ar:
        topic_map = {
            "waiting time": "وقت الانتظار",
            "staff professionalism": "احترافية الموظفين",
            "cleanliness": "النظافة",
            "communication": "التواصل",
            "speed": "سرعة الخدمة",
            "general": "الموضوع العام",
        }
        top_topic_ar = topic_map.get(top_topic, top_topic)
        recent_text = " | ".join(recent) if recent else "لا توجد تعليقات حديثة واضحة."
        return (
            f"بناءً على {len(rows)} ردًا، متوسط التقييم الحالي هو {avg:.1f} من 5. "
            f"أكثر موضوع متكرر الآن هو {top_topic_ar}. "
            f"عدد التعليقات الإيجابية {sentiments.get('Positive', 0)}، والمحايدة {sentiments.get('Neutral', 0)}، والسلبية {sentiments.get('Negative', 0)}. "
            f"إذا كنتِ تبحثين عن خطوة عملية الآن، فالأفضل هو مراجعة التعليقات المتكررة في هذا الموضوع وربطها بالأقسام الأقل تقييمًا. "
            f"أحدث أمثلة من التعليقات: {recent_text}"
        )

    recent_text = " | ".join(recent) if recent else "No recent comments available."
    return (
        f"Based on {len(rows)} responses, the current average rating is {avg:.1f}/5. "
        f"The most repeated topic right now is {top_topic}. "
        f"There are {sentiments.get('Positive', 0)} positive, {sentiments.get('Neutral', 0)} neutral, and {sentiments.get('Negative', 0)} negative comments. "
        f"The most practical next step is to review repeated comments on that topic and compare them with the lowest-rated departments. "
        f"Recent examples: {recent_text}"
    )


def ask_openai(user_message: str, rows: List[sqlite3.Row]) -> str | None:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key or OpenAI is None:
        return None
    client = OpenAI(api_key=api_key)
    context = build_feedback_context(rows)
    response = client.responses.create(
        model=os.getenv("OPENAI_MODEL", "gpt-5.4"),
        instructions=(
            "You are an expert hospital feedback analyst. "
            "Answer in clear English. Base your response only on the provided feedback context. "
            "Be specific, practical, and concise."
        ),
        input=f"Feedback context:\n{context}\n\nUser question: {user_message}",
    )
    return response.output_text.strip()


@app.on_event("startup")
def startup_event() -> None:
    init_db()
    export_excel()


@app.get("/")
def root():
    return {"message": "Feedback backend is running."}


@app.post("/submit")
def submit_feedback(payload: FeedbackIn):
    if not payload.department.strip() or payload.department.strip().lower() == "choose department":
        raise HTTPException(status_code=400, detail="Department is required.")
    if payload.rating1 == 0 and payload.rating2 == 0 and payload.rating3 == 0:
        raise HTTPException(status_code=400, detail="At least one rating is required.")
    avg_rating = round((payload.rating1 + payload.rating2 + payload.rating3) / 3, 2)
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO feedback (created_at, department, section, rating1, rating2, rating3, average_rating, comment) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (created_at, payload.department.strip(), payload.section.strip(), payload.rating1, payload.rating2, payload.rating3, avg_rating, payload.comment.strip()),
        )
        conn.commit()
    export_excel()
    return {"success": True, "message": "Saved successfully.", "excel_file": EXCEL_PATH.name}


@app.get("/summary")
def summary():
    rows = fetch_rows()
    return build_summary_payload(rows)


@app.get("/comments")
def comments():
    rows = fetch_rows(limit=20)
    return {"comments": build_summary_payload(rows).get("latest_comments", [])}


@app.get("/chart")
def chart():
    return {"stars": compute_star_counts(fetch_rows())}


@app.get("/ai")
def ai():
    rows = fetch_rows(limit=50)
    reply = ask_openai("Give me an executive summary of the feedback trends and the next action for management.", rows)
    if not reply:
        reply = fallback_ai_reply("Give me an executive summary of the feedback trends and the next action for management.", rows)
    return {"reply": reply, "using_openai": bool(os.getenv("OPENAI_API_KEY", "").strip() and OpenAI is not None)}


@app.post("/chat")
def chat(payload: ChatIn):
    rows = fetch_rows(limit=80)
    reply = ask_openai(payload.message, rows)
    if not reply:
        reply = fallback_ai_reply(payload.message, rows)
    return {"reply": reply, "using_openai": bool(os.getenv("OPENAI_API_KEY", "").strip() and OpenAI is not None)}


@app.get("/export-excel")
def export_excel_file():
    export_excel()
    return {"success": True, "file": EXCEL_PATH.name}
